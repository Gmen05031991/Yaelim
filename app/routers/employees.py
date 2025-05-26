from fastapi import APIRouter, Request, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import or_
from datetime import date, timedelta
from datetime import datetime
from collections import Counter
from fastapi import HTTPException


from app.models import company as company_model
import shutil, os

from app.database import SessionLocal
from app.models import employee as models
from app.models.document import Document
from app.models.company import Company
from app.models import document as doc_models
from sqlalchemy.orm import relationship
from openpyxl import load_workbook
from fastapi import UploadFile
from io import BytesIO



router = APIRouter(prefix="/employees", tags=["עובדים"])
templates = Jinja2Templates(directory="app/templates")
UPLOAD_DIR = "app/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.get("/", response_class=HTMLResponse)
async def employee_list(
    request: Request,
    db: Session = Depends(get_db),
    q: str = "",
    filter: str = ""
):
    now = datetime.today().date()
    ten_days = now + timedelta(days=10)
    query = db.query(models.Employee)

    # Поиск
    if q:
        query = query.filter(or_(
            models.Employee.first_name.ilike(f"%{q}%"),
            models.Employee.last_name.ilike(f"%{q}%"),
            models.Employee.internal_id.ilike(f"%{q}%")
        ))

    # Обработка фильтра
    if filter == "alpha":
        query = query.order_by(models.Employee.last_name)
    elif filter == "visa":
        query = query.order_by(models.Employee.visa_expiry)
    elif filter == "active":
        query = query.filter(models.Employee.status == "עובד")
    elif filter == "inactive":
        query = query.filter(models.Employee.status == "סיים")
    elif filter.startswith("company_"):
        try:
            company_id = int(filter.split("_")[1])
            query = query.filter(models.Employee.company_id == company_id)
        except ValueError:
            pass

    employees = query.all()

    visa_alerts = db.query(models.Employee).filter(
        models.Employee.visa_expiry <= date.today() + timedelta(days=10),
        models.Employee.status == "עובד"
    ).all()

    companies = db.query(company_model.Company).all()
    

    total_employees = len(employees)
    visa_expiring_soon = len(visa_alerts)
    total_companies = len(companies)
    
    visa_dates = [emp.visa_expiry for emp in employees if emp.visa_expiry]
    visa_date_counts = Counter([date.strftime("%Y-%m-%d") for date in visa_dates])
    visa_chart_labels = list(visa_date_counts.keys())
    visa_chart_data = list(visa_date_counts.values())

    return templates.TemplateResponse("index.html", {
        "total_employees": total_employees,
        "visa_expiring_soon": visa_expiring_soon,
        "total_companies": total_companies,
        "request": request,
        "employees": employees,
        "visa_alerts": visa_alerts,
        "q": q,
        "filter": filter,
        "now": now,
        "ten_days": ten_days,
        "visa_chart_labels": visa_chart_labels,
        "visa_chart_data": visa_chart_data,
       
        "companies": companies
        
        
        
        
    })


@router.get("/{employee_id}/edit", response_class=HTMLResponse)
async def edit_employee_form(employee_id: int, request: Request, db: Session = Depends(get_db)):
    employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    companies = db.query(Company).all()  # ⬅️ ДОБАВЬ ЭТО
    return templates.TemplateResponse("employee_edit.html", {
        "request": request,
        "employee": employee,
        "companies": companies  # ⬅️ И ЭТО
    })

@router.post("/add")
async def add_employee(
     request: Request,
    internal_id: str = Form(...),
    first_name: str = Form(...),
    last_name: str = Form(...),
    phone: str = Form(...),
    birth_date: str = Form(...),  # ← это строка
    passport_number: str = Form(...),
    visa_expiry: str = Form(...),  # ← тоже строка
    address: str = Form(...),
    status: str = Form(...),
    db: Session = Depends(get_db)):

    

    employee = models.Employee(
        internal_id=internal_id,
        first_name=first_name,
        last_name=last_name,
        phone=phone,
        birth_date=birth_date,
        passport_number=passport_number,
        visa_expiry=visa_expiry,
        address=address,
        status=status
    )

     # 🔽 Преобразуем строки в объекты даты
    birth_date_parsed = datetime.strptime(birth_date, "%Y-%m-%d").date()
    visa_expiry_parsed = datetime.strptime(visa_expiry, "%Y-%m-%d").date()

    employee = models.Employee(
        internal_id=internal_id,
        first_name=first_name,
        last_name=last_name,
        phone=phone,
        birth_date=birth_date_parsed,  # 🔽 Передаём как date
        passport_number=passport_number,
        visa_expiry=visa_expiry_parsed,  # 🔽 Передаём как date
        address=address,
        status=status
    )

    db.add(employee)
    db.commit()
    return RedirectResponse("/employees", status_code=303)

@router.get("/{employee_id}", response_class=HTMLResponse)
async def employee_detail(employee_id: int, request: Request, db: Session = Depends(get_db)):
    employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    documents = db.query(doc_models.Document).filter(doc_models.Document.employee_id == employee_id).all()
    return templates.TemplateResponse("employee_card.html", {
        "request": request,
        "employee": employee,
        "documents": documents
    })

@router.get("/{employee_id}/edit", response_class=HTMLResponse)
async def edit_employee_form(employee_id: int, request: Request, db: Session = Depends(get_db)):
    employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    return templates.TemplateResponse("employee_edit.html", {
        "request": request,
        "employee": employee
    })

@router.post("/{employee_id}/edit")
async def update_employee(
    request: Request,
    employee_id: int,
    first_name: str = Form(...),
    last_name: str = Form(...),
    phone: str = Form(...),
    birth_date: str = Form(...),
    passport_number: str = Form(...),
    visa_expiry: str = Form(...),
    address: str = Form(...),
    status: str = Form(...),
    company_id: int = Form(...),
    db: Session = Depends(get_db),
):
    # Парсим даты
    birth_date_parsed = datetime.strptime(birth_date, "%Y-%m-%d").date()
    visa_expiry_parsed = datetime.strptime(visa_expiry, "%Y-%m-%d").date()

    # Находим сотрудника
    employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Обновляем поля
    employee.first_name = first_name
    employee.last_name = last_name
    employee.phone = phone
    employee.birth_date = birth_date_parsed
    employee.passport_number = passport_number
    employee.visa_expiry = visa_expiry_parsed
    employee.address = address
    employee.status = status
    employee.company_id = company_id

    db.commit()
    return RedirectResponse(f"/employees/{employee_id}", status_code=303)

@router.post("/{employee_id}/delete_document/{document_id}")
async def delete_document(employee_id: int, document_id: int, db: Session = Depends(get_db)):
    doc = db.query(doc_models.Document).filter_by(id=document_id, employee_id=employee_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Удалить файл с диска
    file_path = f"app/uploads/{doc.filename}"
    if os.path.exists(file_path):
        os.remove(file_path)

    # Удалить запись из БД
    db.delete(doc)
    db.commit()

    return RedirectResponse(f"/employees/{employee_id}", status_code=303)

@router.post("/{employee_id}/delete")
async def delete_employee(employee_id: int, db: Session = Depends(get_db)):
    employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Удалить документы работника
    documents = db.query(Document).filter(Document.employee_id == employee_id).all()
    for doc in documents:
        file_path = f"{UPLOAD_DIR}/{doc.filename}"
        if os.path.exists(file_path):
            os.remove(file_path)
        db.delete(doc)

    db.delete(employee)
    db.commit()

    return RedirectResponse("/employees", status_code=303)

@router.post("/import_excel")
async def import_employees_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        content = await file.read()
        wb = load_workbook(filename=BytesIO(content))
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            internal_id, first_name, last_name, phone, birth_date, passport_number, visa_expiry, address, status = row[:9]

            # Преобразуем даты
            birth_date = birth_date if isinstance(birth_date, date) else datetime.strptime(birth_date, "%Y-%m-%d").date()
            visa_expiry = visa_expiry if isinstance(visa_expiry, date) else datetime.strptime(visa_expiry, "%Y-%m-%d").date()

            emp = models.Employee(
                internal_id=internal_id,
                first_name=first_name,
                last_name=last_name,
                phone=phone,
                birth_date=birth_date,
                passport_number=passport_number,
                visa_expiry=visa_expiry,
                address=address,
                status=status,
            )
            db.add(emp)

        db.commit()
        return RedirectResponse("/employees", status_code=303)

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"שגיאה בקובץ האקסל: {str(e)}")

@router.post("/{employee_id}/upload")
async def upload_document(employee_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    file_location = f"{UPLOAD_DIR}/{file.filename}"
    with open(file_location, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    doc = doc_models.Document(filename=file.filename, employee_id=employee_id)
    db.add(doc)
    db.commit()
    return RedirectResponse(f"/employees/{employee_id}", status_code=303)
  
  